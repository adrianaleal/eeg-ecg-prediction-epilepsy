'''
************************************* TEST *************************************
Test models (scaler, feature selection, classifier) trained in the training phase

Author: Beatriz Martinho
Date: December 2021
'''
# %% Import functions

# General
import os
import numpy as np
import pandas as pd
import time
import datetime
from datetime import timezone
import fnmatch
# from statistics import mean, pstdev
from scipy import stats

# ML
from import_data import import_feature_data, import_times, import_metadata_info
from build_target import build_target, remove_SPH, remove_postictal
from firing_power import temporal_firing_power_with_decay
from evaluate_performance import did_it_predict_the_seizure, calculate_fpr_denominator  # , calculateFPRdenominatorMauro
from statistical_validation import seizure_surrogate_analysis, surrogate_sensitivity, t_test_one_independent_mean
from utils import load_train_results, get_best_perf
from plot_test_results import plot_each_test_seizure_results, plot_all_test_seizures_results

# Logger
import socket

# Excel writer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# %% Test classifier

# Inputs:
#     - patient_name: string corresponding to the patient identifier in the form 'pat_x'
#                where x is the patient number
#     - approach: string 'Hybrid' or 'Standard'


def test(patient_name: str, signal_type: str, approach: str, perf_metric: str, features_path: str,
         folder2save_figures: str, th_string: str, study_logger):

    study_logger.debug(f'Started testing: {socket.gethostname()}, {patient_name}, {approach}')
    start = time.time()

    # Define patient train results path
    pat_train_results_path = os.path.abspath(os.path.join('Results', signal_type, 'Train_results',
                                                          f'pat_{patient_name}'))

    # Define patient test results path
    pat_test_results_path = os.path.abspath(os.path.join('Results', signal_type, 'Test_results',
                                                         'Test_results_th_' + th_string, f'{patient_name}'))

    # Get the number of training seizures with preictal clustering information
    n_training_seiz_clust_preictal = load_train_results('number_training_seizures_with_clustering_preictal',
                                                        pat_train_results_path)

    # Get best performance
    best_perf = load_train_results(f'{approach}_best_performance', pat_train_results_path)

    # pat_train_results_path_thesis = 'D:\\Results3rdPaper\\ResultsPoster\\ECG\\Train_results\\pat_32702'
    # best_perf_thesis = load_train_results(f'{approach}_best_performance', pat_train_results_path_thesis)
    # best_perf_thesis2 = get_best_perf(approach, pat_train_results_path_thesis)

    # Get parameters corresponding to best performance
    fs_method = best_perf[0]
    mean_preictal_starting_time_min = np.mean(best_perf[1])  # get mean preictal of the training seizures
    Frel = best_perf[3]
    C = best_perf[4]

    # Define firing power threshold
    if th_string == 'opt':
        th = load_train_results(f'{approach}_best_threshold', pat_train_results_path)
        th = th[0]
    else:
        th = float(th_string)

    # Load models
    scaler = load_train_results(f'{approach}_scaler', pat_train_results_path)
    if signal_type == 'ECG':
        fs = load_train_results(f'{approach}_keep_idxs_relevance', pat_train_results_path)
        Frel2 = len(fs)
        f_red = load_train_results(f'{approach}_keep_idxs_redundancy', pat_train_results_path)
        Fred = len(f_red)
        fs_first = Fred
        fs_second = Frel
        name_fs_first = 'Fred'
        name_fs_second = 'Frel'
    else:
        fs = load_train_results(f'{approach}_keep_idxs_redundancy', pat_train_results_path)
        Fred = len(fs)
        fs_first = Frel
        fs_second = Fred
        name_fs_first = 'Frel'
        name_fs_second = 'Fred'
    clf = load_train_results(f'{approach}_clf', pat_train_results_path)

    # Define patient features path
    pat_features_path = os.path.abspath(os.path.join(features_path, f'pat_{patient_name}'))

    # Define number of seizure to train the patient-specific model
    n_training_seizures = 3
    # Get filenames for all seizures in the path
    seizures = fnmatch.filter(os.listdir(pat_features_path), 'feature_datetimes_*.npy')
    n_total_seizures = len(seizures)
    n_testing_seizures = n_total_seizures - n_training_seizures

    # Load training seizure onsets
    metadata_info = import_metadata_info(pat_features_path)

    flag_remove_postictal = 1
    sph_min = 10
    postictal_min = 30
    window_sec = 5
    seizure_sensitivity_all_seizures = 0
    fpr_denominator_all_seizures = 0
    number_false_alarms_all_seizures = 0
    surrogate_sensitivity_rep = []
    # surrogate_sensitivity_bia = []
    nr_repetitions_surrogate = 1000
    seizures_datetimes = []
    seizures_targets = []
    seizures_predicted = []
    seizures_firing_power = []
    seizures_alarms = []

    plot_figure = 1
    if plot_figure == 1:
        from plotly.subplots import make_subplots
        fig = make_subplots(rows=n_testing_seizures, cols=1,
                            subplot_titles=['Patient ' + patient_name + ', Seizure ' + str(seiz_ind + 1)
                                            for seiz_ind in range(n_training_seizures, n_total_seizures)])

    for seiz_ind in range(n_training_seizures, n_total_seizures):

        # print('seizure index', seiz_ind)
        # Load training seizure features
        seizure_features = import_feature_data(pat_features_path, seiz_ind)

        # Load testing times
        seizure_datetimes = import_times(pat_features_path, seiz_ind)

        # Load testing seizure onsets
        seizure_onset = datetime.datetime.fromtimestamp(int(metadata_info[seiz_ind][0]), tz=timezone.utc)

        # Remove SPH from features
        seizure_datetimes, seizure_features = remove_SPH(seizure_datetimes, seizure_onset, sph_min, seizure_features)

        # Remove postictal from features
        if flag_remove_postictal == 1:
            # Load testing seizure offsets
            seizure_offset = datetime.datetime.fromtimestamp(int(metadata_info[seiz_ind][1]), tz=timezone.utc)

            if seiz_ind == n_training_seizures:
                previous_seizure_offset = datetime.datetime.fromtimestamp(int(metadata_info[seiz_ind-1][1]),
                                                                          tz=timezone.utc)

            seizure_datetimes, seizure_features = remove_postictal(seizure_datetimes, previous_seizure_offset,
                                                                   postictal_min, seizure_features)
            previous_seizure_offset = seizure_offset

        # Build targets
        preictal_starting_time_datetime = seizure_onset - datetime.timedelta(minutes=mean_preictal_starting_time_min)
        seizure_target = build_target(seizure_datetimes, preictal_starting_time_datetime)

        if signal_type == 'EEG':
            # Reshape feature arrays (3D to 2D)
            seizure_features = np.reshape(seizure_features, (seizure_features.shape[0],
                                                             seizure_features.shape[1] * seizure_features.shape[2]))

        # Standardization
        seizure_features = scaler.transform(seizure_features)

        if signal_type == 'ECG':
            # # Missing value imputation - find and remove missing values
            # this step was done here instead of doing it in the prepare_data.py, as was done with the EEG, because in
            # ECG there are features that require a minimum number of points to be computed, so it may happen that a
            # given 5-min window could have enough samples for some of the features
            # Train set:
            nan_row_idxs_val = np.unique(np.argwhere(np.isnan(seizure_features))[:, 0])
            # nan_row_idxs_val = np.where(pd.isna(features_tr))
            seizure_features = np.delete(seizure_features, nan_row_idxs_val, axis=0)
            seizure_target = np.delete(seizure_target, nan_row_idxs_val, axis=0)
            seizure_datetimes = np.delete(seizure_datetimes, nan_row_idxs_val, axis=0)

        seizures_datetimes.append(seizure_datetimes)
        seizures_targets.append(seizure_target)

        # Feature selection
        seizure_features = seizure_features[:, [each_fs-1 for each_fs in fs]]

        # Testing
        y_pred = clf.predict(seizure_features)
        seizures_predicted.append(y_pred)

        # Firing Power
        firing_power, predicted_labels = temporal_firing_power_with_decay(y_pred, seizure_datetimes,
                                                                          mean_preictal_starting_time_min, sph_min,
                                                                          window_sec, th)
        seizures_firing_power.append(firing_power)
        seizures_alarms.append(predicted_labels)

        # Plot firing power
        if plot_figure == 1:
            plot_all_test_seizures_results(seiz_ind, n_training_seizures, th, seizure_datetimes, firing_power,
                                           predicted_labels, seizure_target, seizure_onset, fig)

        # Performance Metrics
        seizure_sensitivity = did_it_predict_the_seizure(predicted_labels, seizure_target)
        number_false_alarms, fpr_denominator = calculate_fpr_denominator(predicted_labels, seizure_target,
                                                                         mean_preictal_starting_time_min,
                                                                         seizure_datetimes, seizure_onset, window_sec)
        # fpr_denominator_mauro = calculateFPRdenominatorMauro(predicted_labels, seizure_target,
        #                                                     mean_preictal_starting_time_min, seizure_datetimes,
        #                                                     seizure_onset, window_sec)

        fpr_denominator_all_seizures += fpr_denominator
        number_false_alarms_all_seizures += number_false_alarms
        seizure_sensitivity_all_seizures += 1 * seizure_sensitivity

        # Statistical Validation
        for j in range(nr_repetitions_surrogate):
            surrogate_sensitivity_rep.append(surrogate_sensitivity(predicted_labels, seizure_datetimes, seizure_onset,
                                                                   mean_preictal_starting_time_min, sph_min, th))

        # surrogate_sensitivity_bia = seizure_surrogate_analysis(nr_repetitions_surrogate, seizure_target,
        #                                                        mean_preictal_starting_time_min, predicted_labels,
        #                                                        sph_min, seizure_datetimes, window_sec, th,
        #                                                        surrogate_sensitivity_bia)

    if plot_figure == 1:
        fig.show()
        name2save = patient_name + '_' + signal_type + '_' + approach + '_seizure' + str(seiz_ind + 1)
        fig.write_image(os.path.join(pat_test_results_path, name2save + '.pdf'), engine='kaleido')
        fig.write_html(os.path.join(pat_test_results_path, name2save + '.html'), include_mathjax='cdn')

    fpr_h = number_false_alarms_all_seizures / fpr_denominator_all_seizures
    ss = seizure_sensitivity_all_seizures / n_testing_seizures

    print(f'Sensitivity = {ss}, FPR/h = {fpr_h}')

    mean_surrogate_sensitivity = np.mean(surrogate_sensitivity_rep)
    std_surrogate_sensitivity = np.std(surrogate_sensitivity_rep)
    # 95% confidence interval:
    ci_surrogate_sensitivity = 1.96 * np.std(surrogate_sensitivity_rep) / np.sqrt(len(surrogate_sensitivity_rep))
    print('Surrogate sensitivity (mean +/- std) = ' + str(mean_surrogate_sensitivity) + ' +/- ' +
          str(std_surrogate_sensitivity))
    print('Surrogate sensitivity (mean +/- z*std/sqrt(N)) = ' + str(mean_surrogate_sensitivity) + ' +/- ' +
          str(ci_surrogate_sensitivity))
    # if patient_name == 'pat_98202':  # TODO: check this
    #     seizure_features.pop(1)  # removes the item at the given index from the list
    #     seizure_datetimes.pop(1)
    #     seizure_onsets.pop(1)

    val = 0
    # pval = 1
    pval = float('NaN')
    pval_n = float('NaN')
    tt = float('NaN')
    tt_n = float('NaN')
    print('Does it perform above chance?')
    if mean_surrogate_sensitivity < ss:
        tt_n, pval_n = stats.shapiro(surrogate_sensitivity_rep)
        # if p<=0.05 --> reject the null hypothesis that data was drawn from a normal distribution.
        # One sample t-test is relatively robust to the assumption of normality when the sample size is large (n ≥ 30)
        # https://www.reneshbedre.com/blog/ttest.html

        # [tt, pval] = t_test_one_independent_mean(mean_surrogate_sensitivity, std_surrogate_sensitivity, ss,
        #                                          nr_repetitions_surrogate)

        # Calculate the T-test for the mean of ONE group of scores.
        # Null hypothesis: the expected value (mean) of a sample of independent observations (the 30 surrogate
        # sensitivities), is equal to the hypothesized population mean (ss).
        # Alternative hypothesis:
        # 'two-sided': the mean of the underlying distribution of the sample is different than the given population
        # mean (popmean)
        # 'less': the mean of the underlying distribution of the sample is lesser than the hypothesized population mean
        tt, pval = stats.ttest_1samp(a=surrogate_sensitivity_rep, popmean=ss, alternative='less')

        if pval < 0.05:
            print('Yes')
            val = 1
        else:
            print('No')
    else:
        print('No')

    # # Surrogate Beatriz
    # mean_surrogate_sensitivity_bia = mean(surrogate_sensitivity_bia)
    # std_surrogate_sensitivity_bia = pstdev(surrogate_sensitivity_bia)
    # print('Surrogate sensitivity Beatriz (mean +/- std) = ' + str(mean_surrogate_sensitivity_bia) + ' +/- ' +
    #       str(std_surrogate_sensitivity_bia))
    #
    # if mean_surrogate_sensitivity_bia < ss:
    #     # Perform the Shapiro-Wilk test for normality
    #     # Null hypothesis: data was drawn from a normal distribution
    #     tt_n_bia, pval_n_bia = stats.shapiro(surrogate_sensitivity_bia)
    #     print(f'p-value = {pval_n}')
    #     # If pval_n<0.05 --> reject the null hypothesis of normality
    #
    #     # Calculate the T-test for the mean of ONE group of scores.
    #     # Null hypothesis: the expected value (mean) of a sample of independent observations, surrogate_sensitivity_bia,
    #     # is equal to the given population mean, ss.
    #     # Alternative hypothesis:
    #     # 'two-sided': the mean of the underlying distribution of the sample is different
    #     # than the given population mean (popmean)
    #     # 'less': the mean of the underlying distribution of the sample is less than the given population mean (ss)
    #     tt_bia, p_val = stats.ttest_1samp(surrogate_sensitivity_bia, ss, alternative='less')
    #
    # else:
    #     tt_n_bia = float('NaN')
    #     pval_n_bia = float('NaN')
    #     tt_bia = float('NaN')
    #     pval_bia = float('NaN')

    # Plot predictions
    plot_each_test_seizure_results(patient_name, np.arange(4, 4+n_testing_seizures), seizures_datetimes,
                                   seizures_targets, seizures_predicted, seizures_firing_power, th, seizures_alarms,
                                   fs_method, approach, pat_test_results_path)

    # Save results
    headers = ['Patient', 'No. Train Seizures Clustering Preictal', 'No. Test Seizures', 'Training mean ' + perf_metric,
               'Training std ' + perf_metric, 'Mean Preictal Starting Time', 'Std Preictal Starting Time', 'FS Method',
               name_fs_first, name_fs_second, 'C', 'Threshold', 'Sensitivity', 'FPR/h',
               'Surrogate Sensitivity (Mean +/- Std)', 'Surrogate Sensitivity (Mean +/- z*std/sqrt(N))', 'Shapiro T',
               'Shapiro p-value', 'T-test T', 'T-test p-value', 'Above Chance Flag']
    std_preictal_starting_time_min = np.std(best_perf[1])
    perf_metric_mean = round(best_perf[5], 2)
    perf_metric_std = round(best_perf[6], 2)
    results = [patient_name, sum(n_training_seiz_clust_preictal), n_testing_seizures, perf_metric_mean, perf_metric_std,
               str(round(mean_preictal_starting_time_min, 2)), str(round(std_preictal_starting_time_min, 2)),
               fs_method, fs_first, fs_second, C, th, ss, fpr_h,
               str(round(mean_surrogate_sensitivity, 2)) + u'\u00B1' + str(round(std_surrogate_sensitivity, 2)),
               str(round(mean_surrogate_sensitivity, 2)) + u'\u00B1' + str(round(ci_surrogate_sensitivity, 2)),
               tt_n, pval_n, tt, pval, val]

    save_test_results(patient_name, approach, headers, results, folder2save_figures, signal_type)

    # Compute total runtime
    stop = time.time()
    runtime = round((stop - start), 2)

    study_logger.debug(f'Ended testing: {patient_name}, {approach}, elapsed time: {runtime}')

        
# %% Save Test Results

# Save test results to an excel document. Create a new document or upload existing 
# document. Each document contains the results of each approach for all patients. 

# Inputs:
#     - patient_name: string corresponding to the patient identifier in the form 'pat_x'
#                where x is the patient number
#     - approach: string 'Hybrid' or 'Standard'
#     - headers: list containing the table headers for the results (strings)
#     - results: list containing the results in the same order as the headers


def save_test_results(patient_name: str, approach: str, headers, results, path: str, signal_type: str):

    # If path doesn't exist, create it
    os.makedirs(path, exist_ok=True)
    
    # Define filename and path to file
    file_path = os.path.join(path, f'TestResults' + signal_type + '.xlsx')
    
    # If file exists, open it
    if os.path.exists(file_path):
        
        # Load Workbook object
        wb = openpyxl.load_workbook(file_path)
        
        # Get active sheet
        # ws = wb.active
        if approach not in wb.sheetnames:
            wb.create_sheet(approach)
            ws = wb[approach]
            # Write headers
            for col, val in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = val
                ws.cell(row=1, column=col).font = Font(color='00003366', bold=True)
                ws.cell(row=1, column=col).fill = PatternFill('solid', fgColor='0099CCFF')
                ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col)].width = max(len(val), 10)

            # Get row number to write
            row = ws.max_row + 1

        else:
            ws = wb[approach]
            # Get row number to write
            df = pd.read_excel(file_path, sheet_name=approach)

            get_pat_ind = df.loc[df['Patient'] == int(patient_name)].index.values
            if get_pat_ind.size == 0:
                row = ws.max_row + 1
            else:
                row = get_pat_ind[0] + 2

    # Else, create new file
    else:
        
        # Create new Workbook object
        wb = openpyxl.Workbook()
        
        # Get active sheet
        ws = wb.active
        ws.title = approach

        # Write headers
        for col, val in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = val
            ws.cell(row=1, column=col).font = Font(color='00003366', bold=True)
            ws.cell(row=1, column=col).fill = PatternFill('solid', fgColor='0099CCFF')
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = max(len(val), 10)

        # Get row number to write
        row = ws.max_row + 1

    # Write data into file

    # Write patient results
    for col, val in enumerate(results, start=1):
        ws.cell(row=row, column=col).value = val
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
    # Save file
    wb.save(file_path)
    wb.close()
